import re
import openpyxl
from openpyxl import Workbook

# unrefined_contacts.xlsx
path = 'C:\\Users\\FusRada\\Desktop\\unrefined_contacts.xlsx'


def create_new_excel_file(data_list):
    wb = Workbook()

    ws = wb.active
    ws.title = "Contacts"
    ws['A1'] = "Record ID"
    ws['B1'] = "Names"
    ws['C1'] = "Phone Numbers"

    for x in range(len(data_list)):
        ws['A' + str(x + 2)] = data_list[x]['id']
        ws['B' + str(x + 2)] = data_list[x]['name']
        ws['C' + str(x + 2)] = data_list[x]['number']

    wb.save("refined_contacts.xlsx")


def get_excel_data():
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    data = []

    for row in ws.iter_rows(ws.min_row + 1, ws.max_row):
        row_dict = {'id': None, 'name': None, 'n1': None, 'n2': None}
        for cell in row:
            if cell.column == 1:
                row_dict['id'] = cell.value

            if cell.column == 2:
                row_dict['name'] = cell.value

            if cell.column == 3:
                row_dict['n1'] = cell.value

            if cell.column == 4:
                row_dict['n2'] = cell.value

        data.append(row_dict)

    return data


def extract_numbers(phone_number):
    result = re.sub('[^0-9]', '', phone_number)

    return result


def refine_excel_data(raw_data):
    refined_list = []

    # determine texting number for name, check first for mobile and if not present then use home number
    for x in range(len(raw_data)):
        data_dict = {'id': raw_data[x]['id'], 'name': raw_data[x]['name'], 'number': None}

        if raw_data[x]['n1'] != '':
            value = extract_numbers(raw_data[x]['n1'])
            data_dict['number'] = value
        else:
            value = extract_numbers(raw_data[x]['n2'])
            data_dict['number'] = value

        refined_list.append(data_dict)

    # create list of landlines that need to be removed
    remove_list = []
    for x in range(len(refined_list)):

        if refined_list[x]['number'].__len__() > 11:
            remove_list.append(refined_list[x])

        if refined_list[x]['number'].__len__() < 10:
            remove_list.append(refined_list[x])

    # remove landlines from refined_list
    for x in range(len(remove_list)):
        refined_list.remove(remove_list[x])

    # remove country code that is not USA/Canada
    remove_country = []
    for x in range(len(refined_list)):
        if refined_list[x]['number'].__len__() > 10 and refined_list[x]['number'][:1] != "1":

            # modify 11-digit numbers that start with 0, remove 0
            if refined_list[x]['number'][:1] == "0":
                val = re.sub("0", "", refined_list[x]['number'], 1)
                refined_list[x]['number'] = val
            else:
                remove_country.append(refined_list[x])

    for x in range(len(remove_country)):
        refined_list.remove(remove_country[x])

    final_list = list({v['number']: v for v in refined_list}.values())

    for x in range(len(final_list)):
        first_digit = final_list[x]['number'][0]
        if first_digit == "0":
            print(final_list[x]['number'])

    return final_list


excel_data = get_excel_data()

refined_data = refine_excel_data(excel_data)

print(refined_data.__len__())


def validate_numbers(data):

    for x in range(len(data)):
        num = data[x]['number']
        if len(data[x]['number']) != 10 and len(data[x]['number']) != 11:
            print("Validation Failed: " + str(data[x]))

        if not num.isnumeric():
            print("Validation Failed: " + str(data[x]))


validate_numbers(refined_data)

create_new_excel_file(refined_data)



